"""
Form Monitor v2 - Sistema de monitoreo específico para formulario 5REC
Usa IDs específicos de preguntas y genera tablas detalladas
"""

import pandas as pd
import requests
import json
from datetime import datetime
from pathlib import Path
from fuzzywuzzy import fuzz, process
import time

from config import JOTFORM_API_KEY, FORM_ID

class FormMonitorV2:
    """Monitor específico con IDs de preguntas definidos"""
    
    def __init__(self):
        self.api_key = JOTFORM_API_KEY
        self.form_id = FORM_ID
        self.base_url = "https://api.jotform.com"
        
        # IDs específicos de preguntas
        self.company_field_id = "7"  # Nombre Empresa/Organización
        self.org_type_field_id = "10"  # Tipo de organización
        
        # Data containers
        self.form_questions = {}
        self.all_submissions = []
        self.prefill_records = pd.DataFrame()
        
        print(f"🔍 Form Monitor V2 inicializado")
        print(f"   📋 Formulario ID: {self.form_id}")
        print(f"   🏢 Campo Empresa ID: {self.company_field_id}")
        print(f"   🏛️ Campo Tipo Org ID: {self.org_type_field_id}")

    def load_prefill_records(self):
        """Cargar registros de prefills enviados"""
        print("\n📊 Cargando registros de prefills enviados...")
        
        try:
            outputs_dir = Path("../outputs")
            prefill_files = list(outputs_dir.glob("LINKS_PREFILL_*.xlsx"))
            
            if not prefill_files:
                print("⚠️  No se encontraron archivos de prefill")
                return False
            
            latest_file = max(prefill_files, key=lambda x: x.stat().st_mtime)
            self.prefill_records = pd.read_excel(latest_file, sheet_name="🔗 LINKS PREFILL")
            
            print(f"✅ Prefills cargados: {len(self.prefill_records)} registros")
            print(f"   📁 Archivo: {latest_file.name}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error cargando prefills: {e}")
            return False

    def get_form_structure(self):
        """Obtener estructura completa del formulario"""
        print(f"\n🏗️  Obteniendo estructura del formulario...")
        
        try:
            url = f"{self.base_url}/form/{self.form_id}/questions?apiKey={self.api_key}"
            response = requests.get(url, timeout=30)
            data = response.json()
            
            if response.status_code == 200 and data.get('responseCode') == 200:
                self.form_questions = data.get('content', {})
                
                print(f"✅ Estructura obtenida: {len(self.form_questions)} preguntas")
                
                # Verificar campos específicos
                if self.company_field_id in self.form_questions:
                    company_q = self.form_questions[self.company_field_id]
                    print(f"   🏢 Campo Empresa (ID {self.company_field_id}): {company_q.get('text', 'Sin texto')[:50]}")
                else:
                    print(f"   ⚠️  Campo empresa ID {self.company_field_id} no encontrado")
                
                if self.org_type_field_id in self.form_questions:
                    org_q = self.form_questions[self.org_type_field_id]
                    print(f"   🏛️ Campo Tipo Org (ID {self.org_type_field_id}): {org_q.get('text', 'Sin texto')[:50]}")
                else:
                    print(f"   ⚠️  Campo tipo org ID {self.org_type_field_id} no encontrado")
                
                return True
            else:
                print(f"❌ Error API: {data.get('message', 'Unknown error')}")
                return False
                
        except Exception as e:
            print(f"❌ Error obteniendo formulario: {e}")
            return False

    def get_all_submissions(self):
        """Obtener todas las submissions del formulario"""
        print(f"\n📥 Obteniendo submissions del formulario...")
        
        try:
            url = f"{self.base_url}/form/{self.form_id}/submissions?apiKey={self.api_key}"
            response = requests.get(url, timeout=30)
            data = response.json()
            
            if response.status_code == 200 and data.get('responseCode') == 200:
                self.all_submissions = data.get('content', [])
                
                print(f"✅ Submissions obtenidas: {len(self.all_submissions)}")
                
                if self.all_submissions:
                    for i, submission in enumerate(self.all_submissions[:3]):
                        submission_id = submission.get('id', 'N/A')
                        created_at = submission.get('created_at', 'N/A')
                        print(f"   {i+1}. ID: {submission_id} - Fecha: {created_at}")
                
                return True
            else:
                print(f"❌ Error API: {data.get('message', 'Unknown error')}")
                return False
                
        except Exception as e:
            print(f"❌ Error obteniendo submissions: {e}")
            return False

    def process_submissions(self):
        """Procesar todas las submissions y extraer datos específicos"""
        print(f"\n🔄 Procesando submissions...")
        
        processed_submissions = []
        
        for submission in self.all_submissions:
            answers = submission.get('answers', {})
            
            # Extraer datos específicos
            company_name = self._extract_answer(answers, self.company_field_id, "Sin empresa")
            org_type = self._extract_answer(answers, self.org_type_field_id, "Sin tipo")
            
            # Estado basado en si el formulario está completo o no
            status = submission.get('status', 'ACTIVE')
            estado = "✅ Completado" if status == 'ACTIVE' else "❌ Pendiente"
            
            # Clasificar como prefill o manual
            tipo_submission = self._classify_submission(company_name)
            
            submission_data = {
                'submission_id': submission.get('id'),
                'empresa': company_name,
                'tipo_organizacion': org_type,
                'tipo_submission': tipo_submission['type'],
                'matched_prefill_name': tipo_submission.get('matched_name'),
                'match_score': tipo_submission.get('score', 0),
                'estado': estado,
                'fecha_respuesta': submission.get('created_at'),
                'fecha_actualizacion': submission.get('updated_at'),
                'answers': answers  # Para tabla detallada
            }
            
            processed_submissions.append(submission_data)
        
        print(f"✅ Submissions procesadas: {len(processed_submissions)}")
        
        # Estadísticas
        completed = len([s for s in processed_submissions if s['estado'] == '✅ Completado'])
        prefill_type = len([s for s in processed_submissions if s['tipo_submission'] == 'Prefill'])
        manual_type = len([s for s in processed_submissions if s['tipo_submission'] == 'Manual'])
        
        print(f"   ✅ Completados: {completed}")
        print(f"   📋 Prefill: {prefill_type}")
        print(f"   ✍️  Manual: {manual_type}")
        
        return processed_submissions

    def _extract_answer(self, answers, field_id, default="N/A"):
        """Extraer respuesta específica de un campo"""
        if field_id not in answers:
            return default
        
        answer_data = answers[field_id]
        answer = answer_data.get('answer', default)
        
        if isinstance(answer, list):
            answer = ' '.join(str(item) for item in answer)
        
        return str(answer).strip() if answer else default

    def _is_informational_text(self, text):
        """Identificar si un texto es puramente informativo (no requiere respuesta)"""
        if not text:
            return True
        
        text_lower = text.lower()
        
        # Patrones de textos informativos comunes
        informational_patterns = [
            'a continuación',
            'información adicional',
            'tenga en cuenta',
            'note que',
            'importante:',
            'recuerde que',
            'instrucciones:',
            'por favor',
            'si necesita',
            'para más información',
            'en caso de',
            'si no encuentra',
            'si tiene dudas'
        ]
        
        # Si contiene patrones informativos
        if any(pattern in text_lower for pattern in informational_patterns):
            return True
        
        # Si es muy corto (probablemente un título)
        if len(text.strip()) < 10:
            return True
        
        # Si no termina con signo de pregunta y no parece una solicitud de datos
        if not any(char in text for char in ['?', ':', 'ingrese', 'digite', 'seleccione', 'indique', 'escriba']):
            return True
        
        return False

    def _classify_submission(self, company_name):
        """Clasificar submission como Prefill o Manual"""
        if self.prefill_records.empty or not company_name or company_name == "Sin empresa":
            return {'type': 'Manual', 'matched_name': None, 'score': 0}
        
        # Lista de empresas de prefill
        prefill_companies = self.prefill_records['Empresa/Organización'].tolist()
        
        # Fuzzy matching
        best_match = process.extractOne(
            company_name, 
            prefill_companies,
            scorer=fuzz.token_sort_ratio
        )
        
        if best_match and best_match[1] >= 75:  # 75% similaridad
            return {
                'type': 'Prefill',
                'matched_name': best_match[0],
                'score': best_match[1]
            }
        else:
            return {'type': 'Manual', 'matched_name': None, 'score': 0}

    def add_pending_prefills(self, processed_submissions):
        """Agregar empresas prefill que no han respondido"""
        if self.prefill_records.empty:
            return processed_submissions
        
        # Obtener empresas que ya respondieron
        responded_companies = []
        for submission in processed_submissions:
            if submission['tipo_submission'] == 'Prefill':
                responded_companies.append(submission['matched_prefill_name'])
        
        # Agregar pendientes
        extended_submissions = processed_submissions.copy()
        
        for _, prefill_row in self.prefill_records.iterrows():
            empresa = prefill_row.get('Empresa/Organización', '')
            if empresa and empresa not in responded_companies:
                pending_submission = {
                    'submission_id': 'N/A - Pendiente',
                    'empresa': empresa,
                    'tipo_organizacion': 'N/A - Sin responder',
                    'tipo_submission': 'Prefill',
                    'matched_prefill_name': empresa,
                    'match_score': 100,
                    'estado': '❌ Pendiente',
                    'fecha_respuesta': 'N/A',
                    'fecha_actualizacion': 'N/A',
                    'answers': {}
                }
                extended_submissions.append(pending_submission)
        
        return extended_submissions

    def generate_company_status_table(self, submissions_data, timestamp):
        """Generar tabla 1: Estado por empresa"""
        print(f"\n📊 Generando tabla de estado por empresa...")
        
        table_data = []
        
        for submission in submissions_data:
            table_data.append({
                'Empresa/Organización': submission['empresa'],
                'Tipo Organización': submission['tipo_organizacion'],
                'Tipo Submission': submission['tipo_submission'],
                'Estado': submission['estado'],
                'Fecha Respuesta': submission['fecha_respuesta'],
                'Match Prefill': submission.get('matched_prefill_name', 'N/A'),
                'Score Match': f"{submission.get('match_score', 0)}%",
                'ID Submission': submission['submission_id']
            })
        
        # Crear DataFrame
        df = pd.DataFrame(table_data)
        
        # Ordenar por estado (completados primero) y luego por empresa
        df_sorted = df.sort_values(['Estado', 'Empresa/Organización'], ascending=[False, True])
        
        # Generar Excel
        filename = f"../outputs/ESTADO_EMPRESAS_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_sorted.to_excel(writer, sheet_name='📊 ESTADO EMPRESAS', index=False)
            
            # Formatear
            worksheet = writer.sheets['📊 ESTADO EMPRESAS']
            self._format_status_table(worksheet, len(df_sorted))
        
        print(f"✅ Tabla estado empresas: {filename}")
        return filename

    def generate_detailed_responses_table(self, submissions_data, timestamp):
        """Generar tabla 2: Respuestas detalladas por empresa y pregunta"""
        print(f"\n📋 Generando tabla de respuestas detalladas...")
        
        # Solo procesar submissions que tienen respuestas
        submissions_with_answers = [s for s in submissions_data if s['answers'] and s['estado'] == '✅ Completado']
        
        if not submissions_with_answers:
            print("⚠️  No hay submissions con respuestas para tabla detallada")
            return None
        
        table_data = []
        
        for submission in submissions_with_answers:
            row_data = {
                'Empresa/Organización': submission['empresa'],
                'Tipo Submission': submission['tipo_submission'],
                'Fecha Respuesta': submission['fecha_respuesta'],
                'ID Submission': submission['submission_id']
            }
            
            # Agregar respuesta de cada pregunta (solo campos que requieren respuesta)
            answers = submission['answers']
            for qid, question_info in self.form_questions.items():
                question_text = question_info.get('text', f'Pregunta {qid}')
                question_type = question_info.get('type', 'unknown')
                
                # Excluir campos que NO requieren respuesta
                excluded_types = [
                    'control_head',          # Títulos/encabezados
                    'control_pagebreak',     # Separadores de página
                    'control_button',        # Botones
                    'control_text',          # Textos informativos
                    'control_divider',       # Divisores
                    'control_image',         # Imágenes
                    'control_collapse',      # Secciones colapsables
                    'control_widget'         # Widgets especiales
                ]
                
                if question_type in excluded_types:
                    continue  # Saltar campos informativos
                
                # Filtrar textos que son claramente informativos
                if self._is_informational_text(question_text):
                    continue  # Saltar textos informativos
                
                # Limpiar texto de pregunta (sin cortar)
                clean_question = question_text.replace('\n', ' ').replace('\r', ' ').strip()
                col_name = f"P{qid}: {clean_question}"
                
                # Obtener respuesta
                answer = self._extract_answer(answers, qid, "Sin respuesta")
                row_data[col_name] = answer
            
            table_data.append(row_data)
        
        # Crear DataFrame
        df = pd.DataFrame(table_data)
        
        # Generar Excel
        filename = f"../outputs/RESPUESTAS_DETALLADAS_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='📋 RESPUESTAS DETALLADAS', index=False)
            
            # Formatear
            worksheet = writer.sheets['📋 RESPUESTAS DETALLADAS']
            self._format_detailed_table(worksheet, len(df), len(df.columns))
        
        print(f"✅ Tabla respuestas detalladas: {filename}")
        return filename

    def _format_status_table(self, worksheet, row_count):
        """Formatear tabla de estado"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Anchos de columna
        column_widths = {'A': 40, 'B': 25, 'C': 15, 'D': 15, 'E': 20, 'F': 30, 'G': 12, 'H': 20}
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Freeze panes
        worksheet.freeze_panes = 'A2'

    def _format_detailed_table(self, worksheet, row_count, col_count):
        """Formatear tabla detallada"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Ajustar columnas de forma inteligente
        for col in range(1, col_count + 1):
            col_letter = get_column_letter(col)
            
            if col <= 4:  # Primeras 4 columnas (info básica)
                worksheet.column_dimensions[col_letter].width = 25
            else:  # Columnas de preguntas - ancho adaptativo
                # Calcular ancho basado en el contenido del header
                header_cell = worksheet.cell(row=1, column=col)
                header_text = str(header_cell.value) if header_cell.value else ""
                
                # Ancho mínimo 20, máximo 60, basado en longitud del texto
                header_length = len(header_text)
                if header_length <= 30:
                    width = 25
                elif header_length <= 60:
                    width = 40
                elif header_length <= 100:
                    width = 55
                else:
                    width = 65
                
                worksheet.column_dimensions[col_letter].width = width
        
        # Freeze panes
        worksheet.freeze_panes = 'E2'  # Freeze las primeras 4 columnas
        
        # Altura del header para texto largo
        worksheet.row_dimensions[1].height = 40

    def run_complete_monitoring(self):
        """Ejecutar monitoreo completo"""
        print("=" * 70)
        print("🔍 FORM MONITOR V2 - MONITOREO ESPECÍFICO 5REC")
        print("=" * 70)
        print(f"Formulario ID: {self.form_id}")
        print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Paso 1: Cargar prefills
        has_prefills = self.load_prefill_records()
        
        # Paso 2: Obtener estructura
        if not self.get_form_structure():
            print("❌ No se pudo obtener estructura del formulario")
            return False
        
        # Paso 3: Obtener submissions
        if not self.get_all_submissions():
            print("❌ No se pudieron obtener submissions")
            return False
        
        # Paso 4: Procesar submissions
        processed_submissions = self.process_submissions()
        
        # Paso 5: Agregar prefills pendientes
        all_submissions = self.add_pending_prefills(processed_submissions)
        
        # Paso 6: Generar tablas
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        status_file = self.generate_company_status_table(all_submissions, timestamp)
        detailed_file = self.generate_detailed_responses_table(all_submissions, timestamp)
        
        print("\n" + "=" * 70)
        print("✅ MONITOREO V2 COMPLETADO EXITOSAMENTE")
        print("=" * 70)
        print(f"📊 Total submissions procesadas: {len(processed_submissions)}")
        print(f"📋 Total con prefills pendientes: {len(all_submissions)}")
        print(f"📁 Archivo estado: {Path(status_file).name if status_file else 'N/A'}")
        print(f"📁 Archivo detallado: {Path(detailed_file).name if detailed_file else 'N/A'}")
        print("=" * 70)
        
        return True

def main():
    """Función principal"""
    monitor = FormMonitorV2()
    
    try:
        success = monitor.run_complete_monitoring()
        
        if success:
            print(f"\n🎉 ¡MONITOREO V2 COMPLETADO!")
            print(f"📁 Revisa los archivos Excel en ../outputs/")
        else:
            print(f"\n❌ Problemas en el monitoreo")
            
    except KeyboardInterrupt:
        print("\n\n⏸️  Monitoreo cancelado")
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()