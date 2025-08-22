"""
Excel Generator - Generador de tabla Excel con links de prefill
Sistema que reemplaza el envío de emails por generación de tabla Excel
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import json

class ExcelPrefillGenerator:
    """Generador de tabla Excel con empresas, correos y links de prefill"""
    
    def __init__(self):
        self.results = []
        self.output_dir = Path("../outputs")
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_prefill_excel(self, results_data, form_id=None):
        """
        Genera tabla Excel con empresa, correo y link de prefill
        
        Args:
            results_data: Lista de resultados del motor de prefill
            form_id: ID del formulario (opcional, para metadata)
        """
        print("📊 Generando tabla Excel con links de prefill...")
        
        # Preparar datos para Excel
        excel_data = []
        
        for idx, result in enumerate(results_data, 1):
            empresa = result.get('empresa', f'Empresa_{idx}')
            email = result.get('email', 'N/A')
            prefill_success = result.get('prefill_success', False)
            edit_url = result.get('edit_url', 'N/A')
            mapped_fields = result.get('mapped_fields', 0)
            prefill_error = result.get('prefill_error', '')
            
            # Determinar estado
            if prefill_success:
                estado = "✅ Prefill Exitoso"
                observaciones = f"Campos mapeados: {mapped_fields}"
            else:
                estado = "❌ Error en Prefill"
                observaciones = f"Error: {prefill_error}"
            
            excel_data.append({
                'N°': idx,
                'Empresa/Organización': empresa,
                'Email Destinatario': email,
                'Estado Prefill': estado,
                'Link de Prefill': edit_url if prefill_success else 'N/A',
                'Campos Mapeados': mapped_fields if prefill_success else 0,
                'Observaciones': observaciones,
                'Fecha Procesamiento': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Crear DataFrame
        df = pd.DataFrame(excel_data)
        
        # Generar archivo Excel con formato
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"LINKS_PREFILL_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        # Escribir Excel con formato mejorado
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Hoja principal con datos
            df.to_excel(writer, sheet_name='🔗 LINKS PREFILL', index=False)
            
            # Obtener workbook y worksheet para formateo
            workbook = writer.book
            worksheet = writer.sheets['🔗 LINKS PREFILL']
            
            # Aplicar formato a las columnas
            self._format_excel_worksheet(worksheet, len(df))
            
            # Hoja de resumen
            self._create_summary_sheet(writer, results_data, form_id)
        
        print(f"✅ Tabla Excel generada: {filename}")
        print(f"   📁 Ubicación: {filepath}")
        print(f"   📊 Total registros: {len(excel_data)}")
        
        # Mostrar estadísticas rápidas
        successful_prefills = len([r for r in results_data if r.get('prefill_success', False)])
        print(f"   ✅ Prefills exitosos: {successful_prefills}/{len(results_data)}")
        print(f"   📈 Tasa de éxito: {successful_prefills/len(results_data)*100:.1f}%")
        
        return filepath
    
    def _format_excel_worksheet(self, worksheet, row_count):
        """Aplica formato a la hoja de Excel"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Formato del header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
        
        # Aplicar formato al header (fila 1)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 8,   # N°
            'B': 35,  # Empresa
            'C': 30,  # Email
            'D': 20,  # Estado
            'E': 60,  # Link
            'F': 15,  # Campos
            'G': 40,  # Observaciones
            'H': 20   # Fecha
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Formato para filas de datos
        for row in range(2, row_count + 2):
            # Centrar columnas numéricas
            worksheet[f'A{row}'].alignment = Alignment(horizontal="center")
            worksheet[f'F{row}'].alignment = Alignment(horizontal="center")
            worksheet[f'H{row}'].alignment = Alignment(horizontal="center")
            
            # Wrap text para observaciones
            worksheet[f'G{row}'].alignment = Alignment(wrap_text=True)
        
        # Freeze panes en header
        worksheet.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, writer, results_data, form_id):
        """Crea hoja de resumen con estadísticas"""
        
        # Calcular estadísticas
        total_orgs = len(results_data)
        successful_prefills = len([r for r in results_data if r.get('prefill_success', False)])
        failed_prefills = total_orgs - successful_prefills
        success_rate = (successful_prefills / total_orgs * 100) if total_orgs > 0 else 0
        
        # Datos del resumen
        summary_data = [
            ['RESUMEN PROCESAMIENTO PREFILL 5REC', ''],
            ['', ''],
            ['Parámetro', 'Valor'],
            ['Formulario ID', form_id or 'N/A'],
            ['Fecha Procesamiento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Organizaciones', total_orgs],
            ['Prefills Exitosos', successful_prefills],
            ['Prefills Fallidos', failed_prefills],
            ['Tasa de Éxito (%)', f'{success_rate:.1f}%'],
            ['', ''],
            ['INSTRUCCIONES', ''],
            ['1. Usar columna "Link de Prefill"', 'Copiar y pegar links en emails'],
            ['2. Verificar "Estado Prefill"', 'Solo usar links con estado exitoso'],
            ['3. Personalizar mensaje', 'Incluir nombre de empresa en email'],
            ['4. Monitorear respuestas', 'Seguimiento a formularios completados']
        ]
        
        # Crear DataFrame y escribir
        summary_df = pd.DataFrame(summary_data, columns=['Campo', 'Información'])
        summary_df.to_excel(writer, sheet_name='📋 RESUMEN', index=False)
        
        # Formatear hoja de resumen
        summary_worksheet = writer.sheets['📋 RESUMEN']
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Formato título
        summary_worksheet['A1'].font = Font(bold=True, size=14, color="2C5AA0")
        summary_worksheet.merge_cells('A1:B1')
        
        # Formato header de datos
        summary_worksheet['A3'].font = Font(bold=True)
        summary_worksheet['B3'].font = Font(bold=True)
        
        # Ajustar columnas
        summary_worksheet.column_dimensions['A'].width = 30
        summary_worksheet.column_dimensions['B'].width = 40
        
        # Formato para sección de instrucciones
        for row in range(11, 16):
            summary_worksheet[f'A{row}'].font = Font(bold=True)
    
    def create_email_template(self, empresa, link_prefill, mapped_fields):
        """
        Genera template de email que se puede copiar y pegar
        
        Args:
            empresa: Nombre de la empresa
            link_prefill: URL del formulario prefilled
            mapped_fields: Número de campos mapeados
        """
        
        template = f"""
ASUNTO: Formulario 5REC Pre-llenado - {empresa}

Estimado equipo de {empresa},

Hemos preparado su formulario 5REC con {mapped_fields} campos pre-llenados basados en la información que nos proporcionaron.

🔗 ACCEDER AL FORMULARIO:
{link_prefill}

📝 INSTRUCCIONES:
• El formulario ya tiene información básica de su organización
• Complete los campos faltantes según corresponda  
• Guarde y envíe cuando esté completo
• El enlace es único y personal para su organización

¡Gracias por su participación en el proyecto 5REC!

Equipo 5REC
        """.strip()
        
        return template
    
    def generate_email_templates_file(self, results_data):
        """Genera archivo con templates de email para cada empresa"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"EMAIL_TEMPLATES_{timestamp}.txt"
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("TEMPLATES DE EMAIL PARA FORMULARIOS PREFILL 5REC\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total organizaciones: {len(results_data)}\n\n")
            
            for idx, result in enumerate(results_data, 1):
                if result.get('prefill_success', False):
                    empresa = result.get('empresa', f'Empresa_{idx}')
                    email = result.get('email', 'N/A')
                    link = result.get('edit_url', 'N/A')
                    mapped_fields = result.get('mapped_fields', 0)
                    
                    f.write("-" * 80 + "\n")
                    f.write(f"ORGANIZACIÓN {idx}: {empresa}\n")
                    f.write(f"EMAIL: {email}\n")
                    f.write("-" * 80 + "\n")
                    
                    template = self.create_email_template(empresa, link, mapped_fields)
                    f.write(template)
                    f.write("\n\n")
        
        print(f"✅ Templates de email generados: {filename}")
        return filepath