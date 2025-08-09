"""
Form Analyzer - Advanced JotForm metadata extraction and analysis
Optimized replacement for formMetadata.py with enhanced features
"""

import requests
import json
import csv
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import JOTFORM_API_KEY, FORM_ID

class JotFormAnalyzer:
    """Advanced JotForm metadata analyzer with comprehensive reporting"""
    
    def __init__(self):
        self.api_key = JOTFORM_API_KEY
        self.form_id = FORM_ID
        self.base_url = "https://api.jotform.com"
        self.headers = {"APIKEY": self.api_key}
        self.form_data = {}
        self.questions_data = {}
        self.analysis_results = {}
        
        # Field type translations
        self.field_types_es = {
            'control_textbox': 'Texto simple',
            'control_textarea': 'Área de texto',
            'control_dropdown': 'Lista desplegable',
            'control_radio': 'Opción múltiple (radio)',
            'control_checkbox': 'Casillas de verificación',
            'control_email': 'Email',
            'control_phone': 'Teléfono',
            'control_datetime': 'Fecha y hora',
            'control_number': 'Número',
            'control_fileupload': 'Subir archivo',
            'control_rating': 'Calificación',
            'control_scale': 'Escala',
            'control_matrix': 'Matriz',
            'control_signature': 'Firma',
            'control_head': 'Encabezado/Título',
            'control_pagebreak': 'Salto de página',
            'control_button': 'Botón',
            'control_fullname': 'Nombre completo',
            'control_address': 'Dirección',
            'control_payment': 'Pago',
            'control_captcha': 'Captcha',
            'control_time': 'Hora',
            'control_spinner': 'Selector numérico'
        }
    
    def fetch_form_info(self):
        """Fetch basic form information"""
        print("📋 Obteniendo información del formulario...")
        
        try:
            url = f"{self.base_url}/form/{self.form_id}"
            response = requests.get(url, headers=self.headers, timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                if data.get('responseCode') == 200:
                    self.form_data = data.get('content', {})
                    print(f"✅ Información del formulario obtenida")
                    print(f"   Título: {self.form_data.get('title', 'N/A')}")
                    print(f"   Estado: {self.form_data.get('status', 'N/A')}")
                    return True
            
            print(f"❌ Error obteniendo información: {response.status_code}")
            return False
            
        except Exception as e:
            print(f"❌ Error: {e}")
            return False
    
    def fetch_form_questions(self):
        """Fetch all form questions with detailed analysis"""
        print("\n❓ Obteniendo preguntas del formulario...")
        
        try:
            url = f"{self.base_url}/form/{self.form_id}/questions"
            response = requests.get(url, headers=self.headers, timeout=15)
            
            if response.status_code == 200:
                data = response.json()
                if data.get('responseCode') == 200:
                    self.questions_data = data.get('content', {})
                    print(f"✅ {len(self.questions_data)} preguntas obtenidas")
                    return True
            
            print(f"❌ Error obteniendo preguntas: {response.status_code}")
            return False
            
        except Exception as e:
            print(f"❌ Error: {e}")
            return False
    
    def analyze_form_structure(self):
        """Analyze form structure and generate insights"""
        print("\n🔍 Analizando estructura del formulario...")
        
        if not self.questions_data:
            print("❌ No hay datos de preguntas para analizar")
            return False
        
        # Initialize analysis counters
        field_types = {}
        required_fields = 0
        optional_fields = 0
        fields_with_options = 0
        prefillable_fields = 0
        non_input_fields = 0
        
        # Analyze each question
        for qid, question in self.questions_data.items():
            field_type = question.get('type', 'unknown')
            field_types[field_type] = field_types.get(field_type, 0) + 1
            
            # Check if required
            if question.get('required') == 'Yes':
                required_fields += 1
            else:
                optional_fields += 1
            
            # Check if has options
            if question.get('options'):
                fields_with_options += 1
            
            # Check if prefillable
            if field_type not in ['control_head', 'control_pagebreak', 'control_button', 'control_fileupload']:
                prefillable_fields += 1
            else:
                non_input_fields += 1
        
        # Store analysis results
        self.analysis_results = {
            'total_fields': len(self.questions_data),
            'field_types': field_types,
            'required_fields': required_fields,
            'optional_fields': optional_fields,
            'fields_with_options': fields_with_options,
            'prefillable_fields': prefillable_fields,
            'non_input_fields': non_input_fields
        }
        
        # Print summary
        print(f"✅ Análisis completado:")
        print(f"   • Total de campos: {self.analysis_results['total_fields']}")
        print(f"   • Campos requeridos: {required_fields}")
        print(f"   • Campos opcionales: {optional_fields}")
        print(f"   • Campos prefill-ables: {prefillable_fields}")
        print(f"   • Campos con opciones: {fields_with_options}")
        
        return True
    
    def extract_options(self, options):
        """Extract options from field handling different formats"""
        if not options:
            return ""
        
        try:
            if isinstance(options, dict):
                return "; ".join(str(v) for v in options.values())
            elif isinstance(options, list):
                return "; ".join(str(item) for item in options)
            elif isinstance(options, str):
                return options
            else:
                return str(options)
        except Exception:
            return str(options)
    
    def generate_csv_report(self):
        """Generate comprehensive CSV report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"../outputs/FORM_ANALYSIS_{timestamp}.csv"
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow([
                    'ID_Campo', 'Nombre_Campo', 'Tipo', 'Tipo_ES', 'Requerido',
                    'Orden', 'Opciones', 'Texto_Ayuda', 'Prefillable'
                ])
                
                # Write data
                for qid, question in self.questions_data.items():
                    field_type = question.get('type', '')
                    field_type_es = self.field_types_es.get(field_type, field_type)
                    
                    prefillable = 'Sí' if field_type not in [
                        'control_head', 'control_pagebreak', 'control_button', 'control_fileupload'
                    ] else 'No'
                    
                    writer.writerow([
                        qid,
                        question.get('text', ''),
                        field_type,
                        field_type_es,
                        question.get('required', 'No'),
                        question.get('order', ''),
                        self.extract_options(question.get('options')),
                        question.get('hint', ''),
                        prefillable
                    ])
            
            print(f"✅ Reporte CSV generado: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error generando CSV: {e}")
            return None
    
    def generate_excel_report(self):
        """Generate comprehensive Excel report with multiple sheets"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"../outputs/FORM_ANALYSIS_{timestamp}.xlsx"
        
        try:
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Resumen"
            self._create_summary_sheet(ws_summary)
            
            # Detailed fields sheet
            ws_fields = wb.create_sheet("Campos_Detallados")
            self._create_fields_sheet(ws_fields)
            
            # Analysis sheet
            ws_analysis = wb.create_sheet("Analisis")
            self._create_analysis_sheet(ws_analysis)
            
            # Save workbook
            wb.save(filename)
            print(f"✅ Reporte Excel generado: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error generando Excel: {e}")
            return None
    
    def _create_summary_sheet(self, ws):
        """Create summary sheet with form overview"""
        # Headers
        ws['A1'] = "RESUMEN DEL FORMULARIO"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        
        # Form info
        row = 3
        info_items = [
            ("Título:", self.form_data.get('title', 'N/A')),
            ("ID:", self.form_id),
            ("Estado:", self.form_data.get('status', 'N/A')),
            ("Fecha creación:", self.form_data.get('created_at', 'N/A')),
            ("Total campos:", self.analysis_results.get('total_fields', 0)),
            ("Campos requeridos:", self.analysis_results.get('required_fields', 0)),
            ("Campos prefill-ables:", self.analysis_results.get('prefillable_fields', 0))
        ]
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _create_fields_sheet(self, ws):
        """Create detailed fields sheet"""
        # Headers
        headers = [
            'ID_Campo', 'Nombre_Campo', 'Tipo', 'Tipo_ES', 'Requerido',
            'Orden', 'Opciones', 'Prefillable', 'Texto_Ayuda'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")
        
        # Data
        row = 2
        for qid, question in self.questions_data.items():
            field_type = question.get('type', '')
            field_type_es = self.field_types_es.get(field_type, field_type)
            
            prefillable = 'Sí' if field_type not in [
                'control_head', 'control_pagebreak', 'control_button', 'control_fileupload'
            ] else 'No'
            
            data = [
                qid,
                question.get('text', ''),
                field_type,
                field_type_es,
                question.get('required', 'No'),
                question.get('order', ''),
                self.extract_options(question.get('options')),
                prefillable,
                question.get('hint', '')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_analysis_sheet(self, ws):
        """Create analysis sheet with statistics"""
        ws['A1'] = "ANÁLISIS ESTADÍSTICO"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Field type distribution
        row = 3
        ws[f'A{row}'] = "Distribución por tipo de campo:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for field_type, count in sorted(self.analysis_results['field_types'].items()):
            field_type_es = self.field_types_es.get(field_type, field_type)
            ws[f'A{row}'] = f"{field_type_es}:"
            ws[f'B{row}'] = count
            row += 1
    
    def generate_json_export(self):
        """Generate complete JSON export"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"../outputs/FORM_COMPLETE_{timestamp}.json"
        
        try:
            export_data = {
                'metadata': {
                    'export_timestamp': timestamp,
                    'form_id': self.form_id,
                    'analysis_version': '2.0'
                },
                'form_info': self.form_data,
                'questions': self.questions_data,
                'analysis': self.analysis_results
            }
            
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            print(f"✅ Exportación JSON completa: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error generando JSON: {e}")
            return None
    
    def run_complete_analysis(self):
        """Run complete form analysis workflow"""
        print("=" * 60)
        print("🔬 ANALIZADOR AVANZADO DE FORMULARIO JOTFORM")
        print("=" * 60)
        print(f"Formulario ID: {self.form_id}")
        print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Step 1: Fetch form info
        if not self.fetch_form_info():
            print("\n❌ No se pudo obtener información del formulario")
            return False
        
        # Step 2: Fetch questions
        if not self.fetch_form_questions():
            print("\n❌ No se pudieron obtener las preguntas")
            return False
        
        # Step 3: Analyze structure
        if not self.analyze_form_structure():
            print("\n❌ Error en el análisis de estructura")
            return False
        
        # Step 4: Generate reports
        print("\n📊 Generando reportes...")
        
        csv_file = self.generate_csv_report()
        excel_file = self.generate_excel_report()
        json_file = self.generate_json_export()
        
        print("\n" + "=" * 60)
        print("✅ ANÁLISIS COMPLETADO EXITOSAMENTE")
        
        if csv_file or excel_file or json_file:
            print("\n📁 Archivos generados:")
            if csv_file: print(f"   • CSV: {csv_file}")
            if excel_file: print(f"   • Excel: {excel_file}")
            if json_file: print(f"   • JSON: {json_file}")
        
        print(f"\n➡️  Próximo paso: Usar prefillTemplate_v1.py para generar plantillas")
        print("=" * 60)
        
        return True

def main():
    """Main execution function"""
    analyzer = JotFormAnalyzer()
    
    try:
        success = analyzer.run_complete_analysis()
        
        if not success:
            print("\n🔧 Revisar configuración y conectividad")
            
    except KeyboardInterrupt:
        print("\n\n⏸️  Análisis cancelado por usuario")
    except Exception as e:
        print(f"\n❌ Error inesperado: {e}")
    
    input("\nPresiona Enter para salir...")

if __name__ == "__main__":
    main()